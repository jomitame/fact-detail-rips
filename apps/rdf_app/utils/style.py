from openpyxl.styles import Font, Border, Side, Alignment, PatternFill


def give_style(p_stylo):
    """
    method for give style to a cell

    :param p_stylo: an int that references a style
    :return: a dict with parameters: border, font, align, fill
    """

    my_stylo = {}
    if (p_stylo == 'title'):
        ft = Font(size=8, bold=True, italic=True, color="000000")
        bd = Border(left=Side(border_style="double"), right=Side(border_style="double"),
                    top=Side(border_style="double"), bottom=Side(border_style="double"))
        al = Alignment(horizontal='center', vertical='center')
        rl = PatternFill(fill_type="solid", fgColor="dddddd")
    elif (p_stylo == 'title-wrap'):
        ft = Font(size=8, bold=True, italic=True, color="000000")
        bd = Border(left=Side(border_style="double"), right=Side(border_style="double"),
                    top=Side(border_style="double"), bottom=Side(border_style="double"))
        al = Alignment(horizontal='center', vertical='center', wrap_text=True)
        rl = PatternFill(fill_type="solid", fgColor="dddddd")
    elif (p_stylo == 'encabezado'):
        ft = Font(size=7, bold=True, italic=False, color="000000")
        bd = Border(left=Side(border_style="double"), right=Side(border_style="double"),
                    top=Side(border_style="double"), bottom=Side(border_style="double"))
        al = Alignment(horizontal='center', vertical='center', wrap_text=True)
        rl = PatternFill(fill_type="solid", fgColor="dddddd")
    elif (p_stylo == 'enc-sinwrap'):
        ft = Font(size=8, bold=True, italic=False, color="000000")
        bd = Border(left=Side(border_style="double"), right=Side(border_style="double"),
                    top=Side(border_style="double"), bottom=Side(border_style="double"))
        al = Alignment(horizontal='center', vertical='center', wrap_text=False)
        rl = PatternFill(fill_type="solid", fgColor="dddddd")
    elif (p_stylo == 'total'):
        ft = Font(size=8, bold=False, italic=False, color="000000")
        bd = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                    top=Side(border_style="thin"),
                    bottom=Side(border_style="thin"))
        al = Alignment(horizontal='right', vertical='center')
        rl = PatternFill(fill_type="solid", fgColor="dddddd")
    elif(p_stylo == 'normal'):
        ft = Font(size=7, bold=False, italic=False, color="000000")
        bd = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                    top=Side(border_style="thin"),
                    bottom=Side(border_style="thin"))
        al = Alignment(horizontal='left', vertical='center')
        rl = PatternFill(fill_type="none")
    elif (p_stylo == 'normal-bold'):
        ft = Font(size=7, bold=True, italic=False, color="000000")
        bd = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                    top=Side(border_style="thin"),
                    bottom=Side(border_style="thin"))
        al = Alignment(horizontal='left', vertical='center')
        rl = PatternFill(fill_type="none")
    elif (p_stylo == 'normal-wrap'):
        ft = Font(size=7, bold=False, italic=False, color="000000")
        bd = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                    top=Side(border_style="thin"),
                    bottom=Side(border_style="thin"))
        al = Alignment(horizontal='left', vertical='center', wrap_text=True)
        rl = PatternFill(fill_type="none")
    elif (p_stylo == 'uniblock-top-ef'):
        ft = None
        al = None
        rl = None
        bd = Border(left=Side(border_style="double"), right=Side(border_style="double"),
                    top=Side(border_style="double"), bottom=Side(border_style="none"))
    elif (p_stylo == 'uniblock-top'):
        ft = None
        al = None
        rl = None
        bd = Border(left=Side(border_style="double"), right=Side(border_style="double"),
                    top=Side(border_style="double"), bottom=Side(border_style="thin"))
    elif (p_stylo == 'uniblock-center-ef'):
        ft = None
        al = None
        rl = None
        bd = Border(left=Side(border_style="double"), right=Side(border_style="double"),
                    top=Side(border_style="none"), bottom=Side(border_style="none"))
    elif (p_stylo == 'uniblock-center'):
        ft = None
        al = None
        rl = None
        bd = Border(left=Side(border_style="double"), right=Side(border_style="double"),
                    top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    elif (p_stylo == 'block-left-ef'):
        ft = None
        al = None
        rl = None
        bd = Border(left=Side(border_style="double"), right=Side(border_style="none"),
                    top=Side(border_style="none"), bottom=Side(border_style="none"))
    elif (p_stylo == 'block-left'):
        ft = None
        al = None
        rl = None
        bd = Border(left=Side(border_style="double"), right=Side(border_style="thin"),
                    top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    elif (p_stylo == 'block-right-ef'):
        ft = None
        al = None
        rl = None
        bd = Border(left=Side(border_style="none"), right=Side(border_style="double"),
                    top=Side(border_style="none"), bottom=Side(border_style="none"))
    elif (p_stylo == 'block-right'):
        ft = None
        al = None
        rl = None
        bd = Border(left=Side(border_style="thin"), right=Side(border_style="double"),
                    top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    elif (p_stylo == 'block-btmlft-ef'):
        ft = None
        al = None
        rl = None
        bd = Border(left=Side(border_style="double"), right=Side(border_style="none"),
                    top=Side(border_style="none"), bottom=Side(border_style="double"))
    elif (p_stylo == 'block-btmlft'):
        ft = None
        al = None
        rl = None
        bd = Border(left=Side(border_style="double"), right=Side(border_style="thin"),
                    top=Side(border_style="thin"), bottom=Side(border_style="double"))
    elif (p_stylo == 'block-bottom'):
        ft = None
        al = None
        rl = None
        bd = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                    top=Side(border_style="thin"), bottom=Side(border_style="double"))
    elif (p_stylo == 'block-btmrgt-ef'):
        ft = None
        al = None
        rl = None
        bd = Border(left=Side(border_style="none"), right=Side(border_style="double"),
                    top=Side(border_style="none"), bottom=Side(border_style="double"))
    elif (p_stylo == 'block-btmrgt'):
        ft = None
        al = None
        rl = None
        bd = Border(left=Side(border_style="thin"), right=Side(border_style="double"),
                    top=Side(border_style="thin"), bottom=Side(border_style="double"))

    my_stylo.update({'fuente': ft})
    my_stylo.update({'borde': bd})
    my_stylo.update({'alineacion': al})
    my_stylo.update({'relleno': rl})

    return my_stylo