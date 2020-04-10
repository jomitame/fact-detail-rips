from datetime import datetime
from io import open
import os

from django.shortcuts import render
from openpyxl import Workbook
from django.http import HttpResponse
from django.views.generic import TemplateView
from openpyxl.drawing.image import Image

from apps.rdf_app.models import Fact, DetailMedi, DetailLabo, DetailDispo, DetailService, DetailMediNoPos, Patient
from apps.rdf_app.forms import CreatorForm, GeneratorForm
from apps.rdf_app.utils.style import give_style

class CreatorXLSXView(TemplateView):
    model = Fact
    template_name = 'rdf/creator.html'
    form_class = CreatorForm


    def get(self, request):
        return render(request, self.template_name, {'form':self.form_class})

    def form_invalid(self, form):
        return super(CreatorXLSXView, self).form_invalid(form)

    def fill_detail(self, type, title, mylist, ws, fl, total, *args, **kwargs ):
        if mylist:
            ws.cell(row=fl, column=1).value = title
            ws.cell(row=fl, column=1).fill = give_style('title').get('relleno')
            ws.cell(row=fl, column=1).alignment = give_style('title').get('alineacion')
            ws.cell(row=fl, column=1).font = give_style('title').get('fuente')
            ws.cell(row=fl, column=1).border = give_style('uniblock-top').get('borde')
            ws.merge_cells(start_row=fl, start_column=1, end_row=fl, end_column=8)
            fl += 1
            ws.cell(row=fl, column=1).value = args[0]
            ws.cell(row=fl, column=1).alignment = give_style('enc-sinwrap').get('alineacion')
            ws.cell(row=fl, column=1).font = give_style('encabezado').get('fuente')
            ws.cell(row=fl, column=1).border = give_style('block-left').get('borde')
            ws.cell(row=fl, column=2).value = args[1]
            ws.cell(row=fl, column=2).alignment = give_style('enc-sinwrap').get('alineacion')
            ws.cell(row=fl, column=2).font = give_style('encabezado').get('fuente')
            ws.cell(row=fl, column=2).border = give_style('normal-wrap').get('borde')
            if type==3:
                ws.cell(row=fl, column=3).value = args[7]
                ws.cell(row=fl, column=3).alignment = give_style('enc-sinwrap').get('alineacion')
                ws.cell(row=fl, column=3).font = give_style('encabezado').get('fuente')
                ws.cell(row=fl, column=3).border = give_style('normal-wrap').get('borde')
                ws.cell(row=fl, column=4).value = args[8]
                ws.cell(row=fl, column=4).alignment = give_style('enc-sinwrap').get('alineacion')
                ws.cell(row=fl, column=4).font = give_style('encabezado').get('fuente')
                ws.cell(row=fl, column=4).border = give_style('normal-wrap').get('borde')
            else:
                ws.merge_cells(start_row=fl, start_column=2, end_row=fl, end_column=4)
            ws.cell(row=fl, column=5).value = args[2]
            ws.cell(row=fl, column=5).alignment = give_style('enc-sinwrap').get('alineacion')
            ws.cell(row=fl, column=5).font = give_style('encabezado').get('fuente')
            ws.cell(row=fl, column=5).border = give_style('normal-wrap').get('borde')
            ws.cell(row=fl, column=6).value = args[3]
            ws.cell(row=fl, column=6).alignment = give_style('enc-sinwrap').get('alineacion')
            ws.cell(row=fl, column=6).font = give_style('encabezado').get('fuente')
            ws.cell(row=fl, column=6).border = give_style('normal-wrap').get('borde')
            ws.cell(row=fl, column=7).value = args[4]
            ws.cell(row=fl, column=7).alignment = give_style('enc-sinwrap').get('alineacion')
            ws.cell(row=fl, column=7).font = give_style('encabezado').get('fuente')
            ws.cell(row=fl, column=7).border = give_style('normal-wrap').get('borde')
            ws.cell(row=fl, column=8).value = args[5]
            ws.cell(row=fl, column=8).alignment = give_style('enc-sinwrap').get('alineacion')
            ws.cell(row=fl, column=8).font = give_style('encabezado').get('fuente')
            ws.cell(row=fl, column=8).border = give_style('block-right').get('borde')
            fl += 1
            if type==1:
                for elem in mylist:
                    ws.cell(row=fl, column=1).value = str(getattr(getattr(elem, kwargs['a1']), kwargs['a2']))
                    ws.cell(row=fl, column=1).font = give_style('normal').get('fuente')
                    ws.cell(row=fl, column=1).alignment = give_style('title').get('alineacion')
                    ws.cell(row=fl, column=1).border = give_style('block-left').get('borde')
                    ws.cell(row=fl, column=2).value = str(getattr(getattr(elem, kwargs['a1']), kwargs['a3']))
                    ws.cell(row=fl, column=2).font = give_style('normal').get('fuente')
                    ws.cell(row=fl, column=2).alignment = give_style('normal-wrap').get('alineacion')
                    ws.cell(row=fl, column=2).border = give_style('normal-wrap').get('borde')
                    ws.merge_cells(start_row=fl, start_column=2, end_row=fl, end_column=4)
                    ws.cell(row=fl, column=5).value = str(getattr(elem, kwargs['a4']))
                    ws.cell(row=fl, column=5).font = give_style('normal').get('fuente')
                    ws.cell(row=fl, column=5).alignment = give_style('title').get('alineacion')
                    ws.cell(row=fl, column=5).border = give_style('normal-wrap').get('borde')
                    ws.cell(row=fl, column=6).value = int(getattr(elem, kwargs['a5']))
                    ws.cell(row=fl, column=6).font = give_style('normal').get('fuente')
                    ws.cell(row=fl, column=6).alignment = give_style('title').get('alineacion')
                    ws.cell(row=fl, column=6).border = give_style('normal-wrap').get('borde')
                    ws.cell(row=fl, column=7).value = round(float(getattr(elem, kwargs['a6'])), 2)
                    ws.cell(row=fl, column=7).font = give_style('normal').get('fuente')
                    ws.cell(row=fl, column=7).number_format = '"$"#,##0_);("$"#,##0)'
                    ws.cell(row=fl, column=7).alignment = give_style('total').get('alineacion')
                    ws.cell(row=fl, column=7).border = give_style('normal-wrap').get('borde')
                    ws.cell(row=fl, column=8).value = round(float(getattr(elem, kwargs['a7'])), 2)
                    ws.cell(row=fl, column=8).font = give_style('normal').get('fuente')
                    ws.cell(row=fl, column=8).number_format = '"$"#,##0_);("$"#,##0)'
                    ws.cell(row=fl, column=8).alignment = give_style('total').get('alineacion')
                    ws.cell(row=fl, column=8).border = give_style('block-right').get('borde')
                    fl += 1
            elif type==2:
                for elem in mylist:
                    ws.cell(row=fl, column=1).value = str(kwargs['a7'])
                    ws.cell(row=fl, column=1).font = give_style('normal').get('fuente')
                    ws.cell(row=fl, column=1).alignment = give_style('title').get('alineacion')
                    ws.cell(row=fl, column=1).border = give_style('block-left').get('borde')
                    ws.cell(row=fl, column=2).value = str(getattr(getattr(elem, kwargs['a1']), kwargs['a2']))
                    ws.cell(row=fl, column=2).font = give_style('normal').get('fuente')
                    ws.cell(row=fl, column=2).alignment = give_style('normal-wrap').get('alineacion')
                    ws.cell(row=fl, column=2).border = give_style('normal-wrap').get('borde')
                    ws.merge_cells(start_row=fl, start_column=2, end_row=fl, end_column=4)
                    ws.cell(row=fl, column=5).value = str(getattr(getattr(elem, kwargs['a1']), kwargs['a3']))
                    ws.cell(row=fl, column=5).font = give_style('normal').get('fuente')
                    ws.cell(row=fl, column=5).alignment = give_style('title').get('alineacion')
                    ws.cell(row=fl, column=5).border = give_style('normal-wrap').get('borde')
                    ws.cell(row=fl, column=6).value = int(getattr(elem, kwargs['a4']))
                    ws.cell(row=fl, column=6).font = give_style('normal').get('fuente')
                    ws.cell(row=fl, column=6).alignment = give_style('title').get('alineacion')
                    ws.cell(row=fl, column=6).border = give_style('normal-wrap').get('borde')
                    ws.cell(row=fl, column=7).value = round(float(getattr(elem, kwargs['a5'])), 2)
                    ws.cell(row=fl, column=7).font = give_style('normal').get('fuente')
                    ws.cell(row=fl, column=7).number_format = '"$"#,##0_);("$"#,##0)'
                    ws.cell(row=fl, column=7).alignment = give_style('total').get('alineacion')
                    ws.cell(row=fl, column=7).border = give_style('normal-wrap').get('borde')
                    ws.cell(row=fl, column=8).value = round(float(getattr(elem, kwargs['a6'])), 2)
                    ws.cell(row=fl, column=8).font = give_style('normal').get('fuente')
                    ws.cell(row=fl, column=8).number_format = '"$"#,##0_);("$"#,##0)'
                    ws.cell(row=fl, column=8).alignment = give_style('total').get('alineacion')
                    ws.cell(row=fl, column=8).border = give_style('block-right').get('borde')
                    fl += 1
            elif type==3:
                for elem in mylist:
                    ws.cell(row=fl, column=1).value = str(getattr(getattr(elem, kwargs['a1']), kwargs['a2']))
                    ws.cell(row=fl, column=1).font = give_style('normal').get('fuente')
                    ws.cell(row=fl, column=1).alignment = give_style('title').get('alineacion')
                    ws.cell(row=fl, column=1).border = give_style('block-left').get('borde')
                    ws.cell(row=fl, column=2).value = str(getattr(getattr(elem, kwargs['a1']), kwargs['a3']))
                    ws.cell(row=fl, column=2).font = give_style('normal').get('fuente')
                    ws.cell(row=fl, column=2).alignment = give_style('normal-wrap').get('alineacion')
                    ws.cell(row=fl, column=2).border = give_style('normal-wrap').get('borde')
                    ws.cell(row=fl, column=3).value = '{}'.format(str(getattr(elem, kwargs['a8'])) if
                        getattr(elem, kwargs['a8']) is not None else '-')
                    ws.cell(row=fl, column=3).font = give_style('normal').get('fuente')
                    ws.cell(row=fl, column=3).alignment = give_style('title').get('alineacion')
                    ws.cell(row=fl, column=3).border = give_style('normal-wrap').get('borde')
                    ws.cell(row=fl, column=4).value = '{}'.format(str(getattr(elem, kwargs['a9'])) if
                                                                  getattr(elem, kwargs['a9']) is not None else '-')
                    ws.cell(row=fl, column=4).font = give_style('normal').get('fuente')
                    ws.cell(row=fl, column=4).alignment = give_style('title').get('alineacion')
                    ws.cell(row=fl, column=4).border = give_style('normal-wrap').get('borde')
                    ws.cell(row=fl, column=5).value = str(getattr(elem, kwargs['a4']))
                    ws.cell(row=fl, column=5).font = give_style('normal').get('fuente')
                    ws.cell(row=fl, column=5).alignment = give_style('title').get('alineacion')
                    ws.cell(row=fl, column=5).border = give_style('normal-wrap').get('borde')
                    ws.cell(row=fl, column=6).value = int(getattr(elem, kwargs['a5']))
                    ws.cell(row=fl, column=6).font = give_style('normal').get('fuente')
                    ws.cell(row=fl, column=6).alignment = give_style('title').get('alineacion')
                    ws.cell(row=fl, column=6).border = give_style('normal-wrap').get('borde')
                    ws.cell(row=fl, column=7).value = round(float(getattr(elem, kwargs['a6'])), 2)
                    ws.cell(row=fl, column=7).font = give_style('normal').get('fuente')
                    ws.cell(row=fl, column=7).number_format = '"$"#,##0_);("$"#,##0)'
                    ws.cell(row=fl, column=7).alignment = give_style('total').get('alineacion')
                    ws.cell(row=fl, column=7).border = give_style('normal-wrap').get('borde')
                    ws.cell(row=fl, column=8).value = round(float(getattr(elem, kwargs['a7'])), 2)
                    ws.cell(row=fl, column=8).font = give_style('normal').get('fuente')
                    ws.cell(row=fl, column=8).number_format = '"$"#,##0_);("$"#,##0)'
                    ws.cell(row=fl, column=8).alignment = give_style('total').get('alineacion')
                    ws.cell(row=fl, column=8).border = give_style('block-right').get('borde')
                    fl += 1

            ws.cell(row=fl, column=1).value = args[6]
            ws.cell(row=fl, column=1).fill = give_style('total').get('relleno')
            ws.cell(row=fl, column=1).alignment = give_style('total').get('alineacion')
            ws.cell(row=fl, column=1).font = give_style('total').get('fuente')
            ws.cell(row=fl, column=1).border = give_style('block-btmlft').get('borde')
            ws.merge_cells(start_row=fl, start_column=1, end_row=fl, end_column=7)
            ws.cell(row=fl, column=8).value = round(float(total), 2)
            ws.cell(row=fl, column=8).font = give_style('normal-bold').get('fuente')
            ws.cell(row=fl, column=8).alignment = give_style('total').get('alineacion')
            ws.cell(row=fl, column=8).number_format = '"$"#,##0_);("$"#,##0)'
            ws.cell(row=fl, column=8).border = give_style('block-btmrgt').get('borde')
            fl += 2
        return fl

    def post(self, request):
        my_form = CreatorForm(request.POST)
        if my_form.is_valid():
            workbook = Workbook()
            sheet_temp = workbook.active

            facts = Fact.objects.all()
            for fact in facts:
                # Sheet Creation
                worksheet = workbook.create_sheet(fact.patient.first_name)

                frs_line = 2

                worksheet.cell(row=frs_line, column=1).value = 'DETALLE DE CARGOS DE FACTURA'
                worksheet.cell(row=frs_line, column=1).font = give_style('title').get('fuente')
                worksheet.cell(row=frs_line, column=1).alignment = give_style('title').get('alineacion')
                worksheet.cell(row=frs_line, column=1).border = give_style('uniblock-top-ef').get('borde')
                worksheet.merge_cells(start_row=frs_line, start_column=1, end_row=frs_line, end_column=8)

                frs_line += 1

                worksheet.cell(row=frs_line, column=1).value = 'FACTURA DE VENTA No '+str(fact.cod_fact)
                worksheet.cell(row=frs_line, column=1).font = give_style('title').get('fuente')
                worksheet.cell(row=frs_line, column=1).alignment = give_style('title').get('alineacion')
                worksheet.cell(row=frs_line, column=1).border = give_style('uniblock-center-ef').get('borde')
                worksheet.merge_cells(start_row=frs_line, start_column=1, end_row=frs_line, end_column=8)
                frs_line += 1

                worksheet.cell(row=frs_line, column=1).value = 'Periodo facturado del ' + str(
                    fact.cut_ini) + ' al ' + str(fact.cut_end)
                worksheet.cell(row=frs_line, column=1).font = give_style('title').get('fuente')
                worksheet.cell(row=frs_line, column=1).alignment = give_style('title').get('alineacion')
                worksheet.cell(row=frs_line, column=1).border = give_style('uniblock-center-ef').get('borde')
                worksheet.merge_cells(start_row=frs_line, start_column=1, end_row=frs_line, end_column=8)
                frs_line += 1

                worksheet.cell(row=frs_line, column=1).value = str(fact.regional.company.name) + ' NIT: ' + str(
                    fact.regional.company.number_id) + str(fact.regional.company.cod_verify)
                worksheet.cell(row=frs_line, column=1).font = give_style('title').get('fuente')
                worksheet.cell(row=frs_line, column=1).alignment = give_style('title').get('alineacion')
                worksheet.cell(row=frs_line, column=1).border = give_style('uniblock-center-ef').get('borde')
                worksheet.merge_cells(start_row=frs_line, start_column=1, end_row=frs_line, end_column=8)
                frs_line += 1

                worksheet.cell(row=frs_line, column=1).value = 'NOMBRE'
                worksheet.cell(row=frs_line, column=1).font = give_style('title').get('fuente')
                worksheet.cell(row=frs_line, column=1).border = give_style('block-left-ef').get('borde')
                worksheet.cell(row=frs_line, column=2).value = '{}{} {}{}'.format(str(fact.patient.first_name), (
                            ' ' + str(fact.patient.second_name)) if fact.patient.second_name is not None else '',
                                                     str(fact.patient.first_last_name), (' ' + str(
                        fact.patient.second_last_name)) if fact.patient.second_last_name is not None else '')
                worksheet.cell(row=frs_line, column=2).font = give_style('title').get('fuente')
                worksheet.cell(row=frs_line, column=7).value = 'HC'
                worksheet.cell(row=frs_line, column=7).font = give_style('title').get('fuente')
                worksheet.cell(row=frs_line, column=8).value = str(fact.patient.num_id)
                worksheet.cell(row=frs_line, column=8).font = give_style('title').get('fuente')
                worksheet.cell(row=frs_line, column=8).border = give_style('block-right-ef').get('borde')
                frs_line += 1

                worksheet.cell(row=frs_line, column=1).value = 'DIAGNOSTICO'
                worksheet.cell(row=frs_line, column=1).font = give_style('title').get('fuente')
                worksheet.cell(row=frs_line, column=1).border = give_style('block-left-ef').get('borde')
                worksheet.cell(row=frs_line, column=2).value = str(fact.patient.diagnostic.name)
                worksheet.cell(row=frs_line, column=2).font = give_style('title').get('fuente')
                worksheet.cell(row=frs_line, column=7).value = 'EDAD'
                worksheet.cell(row=frs_line, column=7).font = give_style('title').get('fuente')
                worksheet.cell(row=frs_line, column=8).value = str(fact.patient.age)+' '+str(fact.patient.age_mess)
                worksheet.cell(row=frs_line, column=8).font = give_style('title').get('fuente')
                worksheet.cell(row=frs_line, column=8).border = give_style('block-right-ef').get('borde')
                frs_line += 1

                worksheet.cell(row=frs_line, column=1).value = 'EPS'
                worksheet.cell(row=frs_line, column=1).font = give_style('title').get('fuente')
                worksheet.cell(row=frs_line, column=1).border = give_style('block-btmlft-ef').get('borde')
                worksheet.cell(row=frs_line, column=2).value = str(fact.patient.eps.name)
                worksheet.cell(row=frs_line, column=2).font = give_style('title').get('fuente')
                worksheet.cell(row=frs_line, column=2).alignment = give_style('normal').get('alineacion')
                worksheet.cell(row=frs_line, column=2).border = give_style('block-btmrgt-ef').get('borde')
                worksheet.merge_cells(start_row=frs_line, start_column=2, end_row=frs_line, end_column=8)
                frs_line += 2

                # obtención de datos
                medis_pos = DetailMedi.objects.filter(fact__cod_fact=fact.cod_fact)
                medis_nopos = DetailMediNoPos.objects.filter(fact__cod_fact=fact.cod_fact)
                labos = DetailLabo.objects.filter(fact__cod_fact=fact.cod_fact)
                dispos = DetailDispo.objects.filter(fact__cod_fact=fact.cod_fact)
                servis = DetailService.objects.filter(fact__cod_fact=fact.cod_fact)

                # totales
                total_medis_pos = sum([detalle.subtotal for detalle in
                                       DetailMedi.objects.filter(fact__cod_fact=fact.cod_fact)])

                total_medis_nopos = sum([detalle.subtotal for detalle in
                                         DetailMediNoPos.objects.filter(fact__cod_fact=fact.cod_fact)])
                total_labos = sum(
                    [detalle.subtotal for detalle in DetailLabo.objects.filter(fact__cod_fact=fact.cod_fact)])

                total_dispos = sum(
                    [detalle.subtotal for detalle in DetailDispo.objects.filter(fact__cod_fact=fact.cod_fact)])

                total_services = sum(
                    [detalle.subtotal for detalle in DetailService.objects.filter(fact__cod_fact=fact.cod_fact)])

                '''
                Asi funciona si el valor lo tiene el servicio o producto
                total_services = DetailService.objects.filter(fact__cod_fact=fact.cod_fact).aggregate(
                    sum=Sum(F('cant') * F('service__price'), output_field=FloatField()))
                '''

                # Estancia
                titulos = ['FECHA', 'NOMBRE', 'CODIGO', 'CANTIDAD', 'VALOR UNITARIO', 'VALOR TOTAL',
                           'TOTAL SERVICIO ESTANCIA HOSPITALARIA: ']
                valores = {'a1': 'service', 'a2': 'name', 'a3': 'codigo', 'a4': 'cant', 'a5': 'price',
                           'a6': 'subtotal',
                           'a7': fact.cut_ini}
                frs_line = self.fill_detail(2, 'SERVICIO ESTANCIA HOSPITALARIA', servis,
                                            worksheet,
                                            frs_line,
                                            total_services, *titulos, **valores)
                #frs_line+=1
                # Medicamentos POS
                titulos = ['CODIGO CUM', 'MEDICAMENTOS', 'DOSIS', 'CANTIDAD', 'VALOR UNITARIO', 'VALOR TOTAL',
                           'TOTAL SERVICIO MEDICAMENTOS POS: ']
                valores = {'a1': 'medicine', 'a2': 'cod_cum', 'a3': 'name', 'a4': 'dosis', 'a5': 'cant', 'a6': 'price',
                           'a7': 'subtotal'}
                frs_line = self.fill_detail(1, 'SERVICIO MEDICAMENTOS POS', medis_pos, worksheet, frs_line,
                                 total_medis_pos, *titulos, **valores)

                #frs_line += 1
                # Medicamentos NO POS
                titulos = ['CODIGO CUM', 'MEDICAMENTOS', 'DOSIS', 'CANTIDAD', 'VALOR UNITARIO', 'VALOR TOTAL',
                           'TOTAL SERVICIO MEDICAMENTOS NO POS: ','MIPRES', 'AUTORIZACION']
                valores = {'a1': 'medicine', 'a2': 'cod_cum', 'a3': 'name', 'a4': 'dosis', 'a5': 'cant', 'a6': 'price',
                           'a7': 'subtotal', 'a8':'mipres', 'a9':'autorization'}
                frs_line = self.fill_detail(3, 'SERVICIO MEDICAMENTOS NO POS', medis_nopos, worksheet,
                                            frs_line,
                                            total_medis_nopos, *titulos, **valores)

                #frs_line += 1
                # Dispositivos
                titulos = ['FECHA', 'NOMBRE', 'CODIGO', 'CANTIDAD', 'VALOR UNITARIO', 'VALOR TOTAL',
                           'TOTAL SERVICIO DISPOSITIVOS MEDICO-QUIRURGICOS: ']
                valores = {'a1': 'dispositive', 'a2': 'name', 'a3': 'codigo', 'a4': 'cant', 'a5': 'price',
                           'a6': 'subtotal',
                           'a7': fact.cut_ini}
                frs_line = self.fill_detail(2, 'SERVICIO DISPOSITIVOS MEDICO-QUIRURGICOS', dispos,
                                            worksheet,
                                            frs_line,
                                            total_dispos, *titulos, **valores)

                #frs_line += 1
                # Laboratorios
                titulos = ['FECHA', 'NOMBRE', 'CODIGO', 'CANTIDAD', 'VALOR UNITARIO', 'VALOR TOTAL',
                           'TOTAL SERVICIO LABORATORIOS: ']
                valores = {'a1': 'laboratory', 'a2': 'name', 'a3': 'codigo', 'a4': 'cant', 'a5': 'price', 'a6': 'subtotal',
                           'a7': fact.cut_ini}
                frs_line = self.fill_detail(2, 'SERVICIO LABORATORIO CLINICO', labos,
                                            worksheet,
                                            frs_line,
                                            total_labos, *titulos, **valores)

                #frs_line += 1

                worksheet.cell(row=frs_line, column=1).value = '{}{}{}'.format(
                    'SERVICIO REMEO - PAQUETE/ESTANCIA - ORDEN PQTE: ' + str(fact.aut_number),
                    (' PIN-ELECTRONICO:  ' + str(fact.pin_elect)) if fact.pin_elect is not None else '',
                    (' VALIDACION: ' + str(fact.validation)) if fact.validation is not None else '')
                worksheet.cell(row=frs_line, column=1).fill = give_style('title').get('relleno')
                worksheet.cell(row=frs_line, column=1).alignment = give_style('title-wrap').get('alineacion')
                worksheet.cell(row=frs_line, column=1).font = give_style('title').get('fuente')
                worksheet.cell(row=frs_line, column=1).border = give_style('uniblock-top').get('borde')
                worksheet.merge_cells(start_row=frs_line, start_column=1, end_row=frs_line, end_column=8)
                frs_line += 1
                total_factura = 0
                worksheet.cell(row=frs_line, column=1).value = 'FECHA'
                worksheet.cell(row=frs_line, column=1).alignment = give_style('title').get('alineacion')
                worksheet.cell(row=frs_line, column=1).font = give_style('title').get('fuente')
                worksheet.cell(row=frs_line, column=1).border = give_style('block-left').get('borde')
                worksheet.cell(row=frs_line, column=2).value = 'SERVICIO'
                worksheet.cell(row=frs_line, column=2).alignment = give_style('title').get('alineacion')
                worksheet.cell(row=frs_line, column=2).font = give_style('title').get('fuente')
                worksheet.cell(row=frs_line, column=2).border = give_style('normal-wrap').get('borde')
                worksheet.merge_cells(start_row=frs_line, start_column=2, end_row=frs_line, end_column=7)
                worksheet.cell(row=frs_line, column=8).value = 'TOTAL'
                worksheet.cell(row=frs_line, column=8).alignment = give_style('title').get('alineacion')
                worksheet.cell(row=frs_line, column=8).font = give_style('title').get('fuente')
                worksheet.cell(row=frs_line, column=8).border = give_style('block-right').get('borde')
                frs_line += 1
                if servis:
                    worksheet.cell(row=frs_line, column=1).value = str(fact.cut_ini)
                    worksheet.cell(row=frs_line, column=1).font = give_style('normal').get('fuente')
                    worksheet.cell(row=frs_line, column=1).border = give_style('block-left').get('borde')
                    worksheet.cell(row=frs_line, column=1).alignment = give_style('title').get('alineacion')
                    worksheet.cell(row=frs_line, column=2).value = 'SERVICIO ESTANCIA HOSPITALARIA'
                    worksheet.cell(row=frs_line, column=2).font = give_style('normal').get('fuente')
                    worksheet.cell(row=frs_line, column=2).alignment = give_style('normal').get('alineacion')
                    worksheet.cell(row=frs_line, column=2).border = give_style('normal-wrap').get('borde')
                    worksheet.merge_cells(start_row=frs_line, start_column=2, end_row=frs_line, end_column=7)
                    worksheet.cell(row=frs_line, column=8).value = total_services
                    worksheet.cell(row=frs_line, column=8).font = give_style('total').get('fuente')
                    worksheet.cell(row=frs_line, column=8).number_format = '"$"#,##0_);("$"#,##0)'
                    worksheet.cell(row=frs_line, column=8).border = give_style('block-right').get('borde')
                    worksheet.cell(row=frs_line, column=8).alignment = give_style('total').get('alineacion')
                    total_factura+=total_services
                    frs_line+=1
                if medis_pos:
                    worksheet.cell(row=frs_line, column=1).value = str(fact.cut_ini)
                    worksheet.cell(row=frs_line, column=1).font = give_style('normal').get('fuente')
                    worksheet.cell(row=frs_line, column=1).border = give_style('block-left').get('borde')
                    worksheet.cell(row=frs_line, column=1).alignment = give_style('title').get('alineacion')
                    worksheet.cell(row=frs_line, column=2).value = 'SERVICIO MEDICAMENTOS POS'
                    worksheet.cell(row=frs_line, column=2).font = give_style('normal').get('fuente')
                    worksheet.cell(row=frs_line, column=2).alignment = give_style('normal').get('alineacion')
                    worksheet.cell(row=frs_line, column=2).border = give_style('normal-wrap').get('borde')
                    worksheet.merge_cells(start_row=frs_line, start_column=2, end_row=frs_line, end_column=7)
                    worksheet.cell(row=frs_line, column=8).value = total_medis_pos
                    worksheet.cell(row=frs_line, column=8).font = give_style('total').get('fuente')
                    worksheet.cell(row=frs_line, column=8).number_format = '"$"#,##0_);("$"#,##0)'
                    worksheet.cell(row=frs_line, column=8).border = give_style('block-right').get('borde')
                    worksheet.cell(row=frs_line, column=8).alignment = give_style('total').get('alineacion')
                    total_factura += total_medis_pos
                    frs_line += 1
                if medis_nopos:
                    worksheet.cell(row=frs_line, column=1).value = str(fact.cut_ini)
                    worksheet.cell(row=frs_line, column=1).font = give_style('normal').get('fuente')
                    worksheet.cell(row=frs_line, column=1).border = give_style('block-left').get('borde')
                    worksheet.cell(row=frs_line, column=1).alignment = give_style('title').get('alineacion')
                    worksheet.cell(row=frs_line, column=2).value = 'SERVICIO MEDICAMENTOS NO POS'
                    worksheet.cell(row=frs_line, column=2).font = give_style('normal').get('fuente')
                    worksheet.cell(row=frs_line, column=2).alignment = give_style('normal').get('alineacion')
                    worksheet.cell(row=frs_line, column=2).border = give_style('normal-wrap').get('borde')
                    worksheet.merge_cells(start_row=frs_line, start_column=2, end_row=frs_line, end_column=7)
                    worksheet.cell(row=frs_line, column=8).value = total_medis_nopos
                    worksheet.cell(row=frs_line, column=8).font = give_style('total').get('fuente')
                    worksheet.cell(row=frs_line, column=8).number_format = '"$"#,##0_);("$"#,##0)'
                    worksheet.cell(row=frs_line, column=8).border = give_style('block-right').get('borde')
                    worksheet.cell(row=frs_line, column=8).alignment = give_style('total').get('alineacion')
                    total_factura += total_medis_nopos
                    frs_line += 1
                if dispos:
                    worksheet.cell(row=frs_line, column=1).value = str(fact.cut_ini)
                    worksheet.cell(row=frs_line, column=1).font = give_style('normal').get('fuente')
                    worksheet.cell(row=frs_line, column=1).border = give_style('block-left').get('borde')
                    worksheet.cell(row=frs_line, column=1).alignment = give_style('title').get('alineacion')
                    worksheet.cell(row=frs_line, column=2).value = 'SERVICIO DISPOSITIVOS MEDICO-QUIRURGICOS'
                    worksheet.cell(row=frs_line, column=2).font = give_style('normal').get('fuente')
                    worksheet.cell(row=frs_line, column=2).alignment = give_style('normal').get('alineacion')
                    worksheet.cell(row=frs_line, column=2).border = give_style('normal-wrap').get('borde')
                    worksheet.merge_cells(start_row=frs_line, start_column=2, end_row=frs_line, end_column=7)
                    worksheet.cell(row=frs_line, column=8).value = total_dispos
                    worksheet.cell(row=frs_line, column=8).font = give_style('total').get('fuente')
                    worksheet.cell(row=frs_line, column=8).number_format = '"$"#,##0_);("$"#,##0)'
                    worksheet.cell(row=frs_line, column=8).border = give_style('block-right').get('borde')
                    worksheet.cell(row=frs_line, column=8).alignment = give_style('total').get('alineacion')
                    total_factura += total_dispos
                    frs_line += 1
                if labos:
                    worksheet.cell(row=frs_line, column=1).value = str(fact.cut_ini)
                    worksheet.cell(row=frs_line, column=1).font = give_style('normal').get('fuente')
                    worksheet.cell(row=frs_line, column=1).border = give_style('block-left').get('borde')
                    worksheet.cell(row=frs_line, column=1).alignment = give_style('title').get('alineacion')
                    worksheet.cell(row=frs_line, column=2).value = 'SERVICIO LABORATORIO CLINICO'
                    worksheet.cell(row=frs_line, column=2).font = give_style('normal').get('fuente')
                    worksheet.cell(row=frs_line, column=2).alignment = give_style('normal').get('alineacion')
                    worksheet.cell(row=frs_line, column=2).border = give_style('normal-wrap').get('borde')
                    worksheet.merge_cells(start_row=frs_line, start_column=2, end_row=frs_line, end_column=7)
                    worksheet.cell(row=frs_line, column=8).value = total_labos
                    worksheet.cell(row=frs_line, column=8).font = give_style('total').get('fuente')
                    worksheet.cell(row=frs_line, column=8).number_format = '"$"#,##0_);("$"#,##0)'
                    worksheet.cell(row=frs_line, column=8).border = give_style('block-right').get('borde')
                    worksheet.cell(row=frs_line, column=8).alignment = give_style('total').get('alineacion')
                    total_factura += total_labos
                    frs_line += 1
                worksheet.cell(row=frs_line, column=1).value = 'TOTAL VALOR FACTURADO: '
                worksheet.cell(row=frs_line, column=1).fill = give_style('total').get('relleno')
                worksheet.cell(row=frs_line, column=1).alignment = give_style('total').get('alineacion')
                worksheet.cell(row=frs_line, column=1).font = give_style('total').get('fuente')
                worksheet.cell(row=frs_line, column=1).border = give_style('block-btmlft').get('borde')
                worksheet.merge_cells(start_row=frs_line, start_column=1, end_row=frs_line, end_column=7)
                worksheet.cell(row=frs_line, column=8).value = round(total_factura,2)
                worksheet.cell(row=frs_line, column=8).number_format = '"$"#,##0_);("$"#,##0)'
                worksheet.cell(row=frs_line, column=8).border = give_style('block-btmrgt').get('borde')
                worksheet.cell(row=frs_line, column=8).font = give_style('enc-sinwrap').get('fuente')
                worksheet.cell(row=frs_line, column=8).alignment = give_style('total').get('alineacion')

                worksheet.column_dimensions['A'].width = 12
                worksheet.column_dimensions['B'].width = 20
                worksheet.column_dimensions['C'].width = 9
                worksheet.column_dimensions['D'].width = 12
                worksheet.column_dimensions['E'].width = 7
                worksheet.column_dimensions['F'].width = 7
                worksheet.column_dimensions['G'].width = 12
                worksheet.column_dimensions['H'].width = 9

                img = Image("messer-logo.png")
                img.width = 120
                img.height = 60
                worksheet.add_image(img, 'A2')


            workbook.remove(sheet_temp)


            response = HttpResponse(
                content_type = 'application/ms-excel'
            )
            response['Content-Disposition'] = 'attachment; filename = detallado_{date}.xlsx'.format(
                date=datetime.now().strftime("%d%m%Y-%H%M%S"),
            )
            workbook.save(response)
            return response
        return render(request, "rdf/creator.html", {'form': my_form})

class GeneratorRIPSView(TemplateView):
    model = Fact
    template_name = 'rdf/generator.html'
    form_class = GeneratorForm

    def create_AF_file(self, dir, cod_hab, nomb_emp, type_id_emp, num_id_emp, factu, f_ini, f_end, cod_eps, nomb_eps, total):
        f_pa_dir = f_end.split("/")
        f_pa_dir[-1] = f_pa_dir[-1][-2:]
        f_pa_dir = "".join(f_pa_dir)
        ruta = dir+"/AF{}.txt".format(f_pa_dir)
        os.makedirs(dir, exist_ok=True)
        with open (ruta, "a") as f:
            f.write(
                '{},{},{},{},{},{},{},{},{},{},0,0,0,0,0,0,{}\n'.format(cod_hab, nomb_emp, type_id_emp, num_id_emp, factu,
                                                                     f_end, f_ini, f_end, cod_eps, nomb_eps, total)
            )
            f.close()

    def create_AH_file(self, dir, factu, cod_hab, type_id_pac, num_id_pac, f_ini, num_auto, dx,f_end):
        f_pa_dir = f_end.split("/")
        f_pa_dir[-1] = f_pa_dir[-1][-2:]
        f_pa_dir = "".join(f_pa_dir)
        ruta = dir+"/AH{}.txt".format(f_pa_dir)
        os.makedirs(dir, exist_ok=True)
        with open (ruta, "a") as f:
            f.write(
                '{},{},{},{},3,{},16:05,{},13,{},{},{},{},{},{},1,,{},23:59\n'.format(factu, cod_hab, type_id_pac,
                                                                                      num_id_pac, f_ini, num_auto, dx,
                                                                                      dx, dx, dx, dx, dx, f_end)
            )
            f.close()

    def create_AM_file(self, dir, f_end, factu, cod_hab, type_id_pac, num_id_pac, num_auto, cum, posnopos, nom_mdto,
                       pre_mdto, can_con_mdto, conc_mdto, cant, val_uni, subtotal):
        f_pa_dir = f_end.split("/")
        f_pa_dir[-1] = f_pa_dir[-1][-2:]
        f_pa_dir = "".join(f_pa_dir)
        ruta = dir + "/AM{}.txt".format(f_pa_dir)
        os.makedirs(dir, exist_ok=True)
        with open(ruta, "a") as f:
            f.write(
                '{},{},{},{},{},{},{},{},{},{},{},{},{},{}\n'.format(factu, cod_hab, type_id_pac, num_id_pac, num_auto,
                                                                     cum, posnopos, nom_mdto, pre_mdto, can_con_mdto,
                                                                     conc_mdto, cant, val_uni, subtotal)

            )
            f.close()

    def create_AP_file(self, dir, f_end, factu, cod_hab, type_id_pac, num_id_pac, f_ini, num_auto, cod_proc, dx, subtotal):
        f_pa_dir = f_end.split("/")
        f_pa_dir[-1] = f_pa_dir[-1][-2:]
        f_pa_dir = "".join(f_pa_dir)
        ruta = dir + "/AP{}.txt".format(f_pa_dir)
        os.makedirs(dir, exist_ok=True)
        with open(ruta, "a") as f:
            f.write(
                '{},{},{},{},{},{},{},2,1,5,{},,,,{}\n'.format(factu, cod_hab, type_id_pac, num_id_pac, f_ini, num_auto,
                                                               cod_proc, dx, subtotal)

            )
            f.close()

    def create_AT_file(self, dir, f_end, factu, cod_hab, type_id_pac, num_id_pac, num_auto, cod_ser, nom_ser, cant,
                       val_uni, subtotal):
        f_pa_dir = f_end.split("/")
        f_pa_dir[-1] = f_pa_dir[-1][-2:]
        f_pa_dir = "".join(f_pa_dir)
        ruta = dir + "/AT{}.txt".format(f_pa_dir)
        os.makedirs(dir, exist_ok=True)
        with open(ruta, "a") as f:
            f.write(
                '{},{},{},{},{},3,{},{},{},{},{}\n'.format(factu, cod_hab, type_id_pac, num_id_pac, num_auto, cod_ser,
                                                           nom_ser, cant, val_uni, subtotal)
            )
            f.close()

    def create_US_file(self, dir, f_end, type_id_pac, num_id_pac, cod_eps, regi, a1, a2, n1, n2, age, mess, gene, dpto, muni, rulurb):
        f_pa_dir = f_end.split("/")
        f_pa_dir[-1] = f_pa_dir[-1][-2:]
        f_pa_dir = "".join(f_pa_dir)
        ruta = dir + "/US{}.txt".format(f_pa_dir)
        os.makedirs(dir, exist_ok=True)
        with open(ruta, "a") as f:
            f.write(
                '{},{},{},{},{},{},{},{},{},{},{},{},{},{}\n'.format(type_id_pac, num_id_pac, cod_eps, regi, a1, a2, n1,
                                                                     n2, age, mess, gene, dpto, muni, rulurb)
            )
            f.close()

    def create_CT_file(self, dir, f_end, num_hab_emp, file, cant_lines):
        f_pa_dir = f_end.split("/")
        f_pa_dir[-1] = f_pa_dir[-1][-2:]
        f_pa_dir = "".join(f_pa_dir)
        ruta = dir + "/CT{}.txt".format(f_pa_dir)
        os.makedirs(dir, exist_ok=True)
        with open(ruta, "a") as f:
            f.write(
                '{},{},{},{}\n'.format(num_hab_emp, f_end, file, cant_lines)
            )
            f.close()

    def ct_caller(self, dirs_pa_ct, num_hab_emp, f_end):
        f_pa_dir = f_end.split("/")
        f_pa_dir[-1] = f_pa_dir[-1][-2:]
        f_pa_dir = "".join(f_pa_dir)
        for dir in dirs_pa_ct:
            for prefix in ['AF','AH','AM', 'AP','AT','US']:
                file = "{}{}".format(prefix,f_pa_dir)
                ruta = dir + "/" + file +".txt"
                try:
                    fichero = open(ruta, 'r')
                    fichero.readline()
                    fichero.seek(0)
                    cant_lines = len(fichero.readlines())
                    fichero.close()
                    self.create_CT_file(dir, f_end, num_hab_emp, file, cant_lines)
                except:
                    pass



    def post(self, request):
        my_form = GeneratorForm(request.POST)
        #if my_form.is_valid():
        facts = Fact.objects.all()
        patients = []
        # ----------------------------------------------------------
        dirs_pa_ct = []
        num_id_emp = 0
        f_end = ""

        for fact in facts:

            # obtención de datos
            medis_pos = DetailMedi.objects.filter(fact__cod_fact=fact.cod_fact)
            medis_nopos = DetailMediNoPos.objects.filter(fact__cod_fact=fact.cod_fact)
            labos = DetailLabo.objects.filter(fact__cod_fact=fact.cod_fact)
            dispos = DetailDispo.objects.filter(fact__cod_fact=fact.cod_fact)
            servis = DetailService.objects.filter(fact__cod_fact=fact.cod_fact)

            # totales
            total_medis_pos = sum([detalle.subtotal for detalle in
                                   DetailMedi.objects.filter(fact__cod_fact=fact.cod_fact)])

            total_medis_nopos = sum([detalle.subtotal for detalle in
                                     DetailMediNoPos.objects.filter(fact__cod_fact=fact.cod_fact)])
            total_labos = sum(
                [detalle.subtotal for detalle in DetailLabo.objects.filter(fact__cod_fact=fact.cod_fact)])

            total_dispos = sum(
                [detalle.subtotal for detalle in DetailDispo.objects.filter(fact__cod_fact=fact.cod_fact)])

            total_services = sum(
                [detalle.subtotal for detalle in DetailService.objects.filter(fact__cod_fact=fact.cod_fact)])

            eps = fact.patient.eps.name_rips
            dir = eps+"/"+fact.regional.name_rips
            if dir not in dirs_pa_ct:
                dirs_pa_ct.append(dir)
            cod_hab = fact.regional.cod_hab
            nomb_emp = fact.regional.company.name_rips
            type_id_emp = fact.regional.company.type_id
            num_id_emp = fact.regional.company.number_id
            factu = fact.cod_fact
            f_ini = fact.cut_ini.strftime("%d/%m/%Y")
            f_end = fact.cut_end.strftime("%d/%m/%Y")
            cod_eps = fact.patient.eps.cod_eps
            type_id_pac = fact.patient.type_id
            num_id_pac = fact.patient.num_id
            num_auto = fact.aut_number
            dx = fact.patient.diagnostic.cod_dx

            total = 0


            if fact.patient.num_id not in patients:
                patients.append(fact.patient.num_id)
                regi = fact.patient.regimen
                a1 = fact.patient.first_last_name
                a2 = fact.patient.second_last_name if fact.patient.second_last_name is not None else ""
                n1 = fact.patient.first_name
                n2 = fact.patient.second_name if fact.patient.second_name is not None else ""
                age = fact.patient.age
                mess = '1' if fact.patient.age_mess == 'años' else ('2' if fact.patient.age_mess == 'meses' else'3')
                gene = fact.patient.gene
                dpto = fact.regional.dpto.codigo
                muni = fact.regional.municipe.codigo
                rulurb = fact.regional.urba_rul
                self.create_US_file(dir, f_end, type_id_pac, num_id_pac, cod_eps, regi, a1, a2, n1, n2, age,
                                        mess, gene, dpto,muni, rulurb)

            if servis:
                total += total_services
                self.create_AH_file(dir, factu, cod_hab, type_id_pac, num_id_pac, f_ini, num_auto, dx, f_end)
                for servi in servis:
                    cod_ser = servi.service.codigo
                    nom_ser = servi.service.name_rips
                    cant = servi.cant
                    val_uni = int(servi.price)
                    subtotal = int(servi.subtotal)
                    self.create_AT_file(dir, f_end, factu, cod_hab, type_id_pac, num_id_pac, num_auto, cod_ser,
                                   nom_ser, cant, val_uni, subtotal)

            if medis_pos:
                total += total_medis_pos
                for medi in medis_pos:
                    cum = medi.medicine.cod_cum
                    nom_mdto = medi.medicine.name_rips
                    pre_mdto = medi.medicine.presentation.name_rips
                    can_con_mdto = int(medi.medicine.cant_concent)
                    conc_mdto = medi.medicine.concentration.name_rips
                    cant = int(medi.cant)
                    val_uni = int(medi.price)
                    subtotal = int(medi.subtotal)
                    self.create_AM_file(dir, f_end, factu, cod_hab, type_id_pac, num_id_pac, num_auto, cum, 1,
                                   nom_mdto, pre_mdto, can_con_mdto, conc_mdto, cant, val_uni, subtotal)


            if labos:
                total += total_labos
                for labo in labos:
                    cod_proc = labo.laboratory.codigo
                    subtotal = int(labo.subtotal)
                    self.create_AP_file(dir, f_end, factu, cod_hab, type_id_pac, num_id_pac, f_ini, num_auto, cod_proc, dx,
                               subtotal)

            if dispos:
                total += total_dispos
                for dispo in dispos:
                    cod_proc = dispo.dispositive.codigo
                    subtotal = int(dispo.subtotal)
                    self.create_AP_file(dir, f_end, factu, cod_hab, type_id_pac, num_id_pac, f_ini, num_auto, cod_proc,
                                        dx, subtotal)


            if medis_nopos:
                total += total_medis_nopos
                for medi in medis_nopos:
                    cum = medi.medicine.cod_cum
                    nom_mdto = medi.medicine.name_rips
                    num_auto = medi.autorization# ------------------------------------------------->>>>>>>>>>>>
                    pre_mdto = medi.medicine.presentation.name_rips
                    can_con_mdto = int(medi.medicine.cant_concent)
                    conc_mdto = medi.medicine.concentration.name_rips
                    cant = int(medi.cant)
                    val_uni = int(medi.price)
                    subtotal = int(medi.subtotal)
                    self.create_AM_file(dir, f_end, factu, cod_hab, type_id_pac, num_id_pac, num_auto, cum, 2,
                                   nom_mdto, pre_mdto, can_con_mdto, conc_mdto, cant, val_uni, subtotal)


            total = int(total)
            self.create_AF_file(dir, cod_hab, nomb_emp, type_id_emp, num_id_emp, factu, f_ini, f_end, cod_eps, eps,total)

        self.ct_caller(dirs_pa_ct, cod_hab, f_end)



        return render(request, "rdf/generator.html", {'form': my_form})
